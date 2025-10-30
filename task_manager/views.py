from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.views.decorators.http import require_http_methods
from django.db.models import Count, Q
from datetime import datetime
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from .models import Task


def task_list(request):
    tasks = Task.objects.all()
    return render(request, 'task_manager/task_list.html', {'tasks': tasks})


def task_detail(request, pk):
    task = get_object_or_404(Task, pk=pk)
    return render(request, 'task_manager/partials/task_item.html', {'task': task})


def task_create_modal(request):
    return render(request, 'task_manager/partials/modal_create_task.html')


@require_http_methods(["POST"])
def task_create(request):
    title = request.POST.get('title')
    description = request.POST.get('description', '')

    if title:
        task = Task.objects.create(title=title, description=description)
        return render(request, 'task_manager/partials/task_item.html', {'task': task})

    return HttpResponse(status=400)


@require_http_methods(["GET", "POST"])
def task_update(request, pk):
    task = get_object_or_404(Task, pk=pk)

    if request.method == 'POST':
        task.title = request.POST.get('title', task.title)
        task.description = request.POST.get('description', task.description)
        task.save()
        return render(request, 'task_manager/partials/task_item.html', {'task': task})

    return render(request, 'task_manager/partials/task_form.html', {'task': task})


@require_http_methods(["POST"])
def task_toggle(request, pk):
    task = get_object_or_404(Task, pk=pk)
    task.completed = not task.completed
    task.save()
    return render(request, 'task_manager/partials/task_item.html', {'task': task})


@require_http_methods(["DELETE"])
def task_delete(request, pk):
    task = get_object_or_404(Task, pk=pk)
    task.delete()
    return HttpResponse(status=200)


def export_tasks_report(request):
    tasks = Task.objects.all()
    total_tasks = tasks.count()
    completed_tasks = tasks.filter(completed=True).count()
    pending_tasks = tasks.filter(completed=False).count()
    completion_rate = (completed_tasks / total_tasks * 100) if total_tasks > 0 else 0

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="reporte_tareas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'

    writer = csv.writer(response)

    writer.writerow(['REPORTE DE TAREAS - TASK MANAGER BESIMPLIT'])
    writer.writerow([f'Fecha de generación: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'])
    writer.writerow([])

    writer.writerow(['RESUMEN EJECUTIVO'])
    writer.writerow(['Métrica', 'Valor'])
    writer.writerow(['Total de tareas creadas', total_tasks])
    writer.writerow(['Tareas completadas', completed_tasks])
    writer.writerow(['Tareas pendientes', pending_tasks])
    writer.writerow(['Porcentaje de completitud', f'{completion_rate:.1f}%'])
    writer.writerow([])

    writer.writerow(['DETALLE DE TAREAS'])
    writer.writerow(['ID', 'Título', 'Descripción', 'Estado', 'Fecha Creación', 'Última Actualización'])

    for task in tasks:
        writer.writerow([
            task.id,
            task.title,
            task.description,
            'Completada' if task.completed else 'Pendiente',
            task.created_at.strftime('%d/%m/%Y %H:%M:%S'),
            task.updated_at.strftime('%d/%m/%Y %H:%M:%S')
        ])

    return response


def export_tasks_excel(request):
    tasks = Task.objects.all()
    total_tasks = tasks.count()
    completed_tasks = tasks.filter(completed=True).count()
    pending_tasks = tasks.filter(completed=False).count()
    completion_rate = (completed_tasks / total_tasks * 100) if total_tasks > 0 else 0

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Tareas"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    center_align = Alignment(horizontal="center", vertical="center")

    ws['A1'] = 'REPORTE DE TAREAS - TASK MANAGER BESIMPLIT'
    ws['A1'].font = title_font
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = center_align

    ws['A2'] = f'Fecha de generación: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'
    ws.merge_cells('A2:F2')
    ws['A2'].alignment = center_align

    ws['A4'] = 'RESUMEN EJECUTIVO'
    ws['A4'].font = Font(bold=True, size=12)
    ws.merge_cells('A4:B4')

    headers = ['Métrica', 'Valor']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    summary_data = [
        ['Total de tareas creadas', total_tasks],
        ['Tareas completadas', completed_tasks],
        ['Tareas pendientes', pending_tasks],
        ['Porcentaje de completitud', f'{completion_rate:.1f}%']
    ]

    for row_num, row_data in enumerate(summary_data, 6):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    detail_row = len(summary_data) + 7
    ws[f'A{detail_row}'] = 'DETALLE DE TAREAS'
    ws[f'A{detail_row}'].font = Font(bold=True, size=12)
    ws.merge_cells(f'A{detail_row}:F{detail_row}')

    detail_headers = ['ID', 'Título', 'Descripción', 'Estado', 'Fecha Creación', 'Última Actualización']
    header_row = detail_row + 1
    for col_num, header in enumerate(detail_headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for task_num, task in enumerate(tasks, header_row + 1):
        ws.cell(row=task_num, column=1, value=task.id)
        ws.cell(row=task_num, column=2, value=task.title)
        ws.cell(row=task_num, column=3, value=task.description)
        ws.cell(row=task_num, column=4, value='Completada' if task.completed else 'Pendiente')
        ws.cell(row=task_num, column=5, value=task.created_at.strftime('%d/%m/%Y %H:%M:%S'))
        ws.cell(row=task_num, column=6, value=task.updated_at.strftime('%d/%m/%Y %H:%M:%S'))

    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 20

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_tareas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

    wb.save(response)
    return response
